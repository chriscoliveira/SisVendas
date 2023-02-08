import sqlite3
from time import sleep
import os
import re
import sys
from datetime import datetime
import calendar
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

sistema = sys.platform
if sistema == 'linux':
    nome_cupom = 'CUPOM/cupom_'
else:
    nome_cupom = 'CUPOM\\cupom_'


class Operacoes:

    def __init__(self, arquivo):
        self.conn = sqlite3.connect(arquivo)
        self.cursor = self.conn.cursor()

    def cadastrarProduto(self, ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra, botao):
        # print(botao)
        try:
            if ativo == 'SIM':
                ativo = '1'
            else:
                ativo = '2'

            ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra = str(ativo).upper(), str(ean).upper(), str(
                produto).upper(), str(quantidade).upper(), str(valor_custo).upper().replace(',', '.'), str(valor_venda).upper().replace(',', '.'), str(ultima_compra).upper()

            rativo, rean, rproduto, rquantidade, rvalor_custo, rvalor_venda, rultima_compra = self.buscarProduto(
                termo=ean)

            if str(botao) == 'Cadastrar':
                if rean == ean:
                    return False
                else:

                    sql = "INSERT INTO produtos (ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra) VALUES (?,?,?,?,?,?,?)"
                    self.cursor.execute(
                        sql, (ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra))
                    self.conn.commit()
                    return True
            else:

                try:
                    sql = "UPDATE produtos SET ativo=?, produto=?, quantidade=?, valor_custo=?, valor_venda=? WHERE ean=?"
                    self.cursor.execute(sql, (ativo, produto, quantidade,
                                        valor_custo, valor_venda, ean))
                    self.conn.commit()
                    return True
                except Exception as e:
                    # print(e)
                    return False

        except Exception as e:
            # print(e)
            return False

    def editar(self, id, ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra):
        try:
            ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra = str(ativo).upper(), str(ean).upper(), str(
                produto).upper(), str(quantidade).upper(), str(valor_custo).upper().replace(',', '.'), str(valor_venda).upper().replace(',', '.'), str(ultima_compra).upper()
            sql = "UPDATE produtos SET ativo=?, ean=?, produto=?, quantidade=?, valor_custo=?, valor_venda=?, ultima_compra=? WHERE id=?"
            self.cursor.execute(sql, (ativo, ean, produto, quantidade,
                                valor_custo, valor_venda, ultima_compra, id))
            self.conn.commit()
            return f"O {produto=} {ean=} foi alterado com sucesso!"
        except Exception as e:
            return f"Ocorreu um erro ao alterar = {e}"

    def listar_tudo(self, tabela, tudo=False):
        itens = []
        if tudo:
            sql = "SELECT * FROM "+tabela
        else:
            sql = "SELECT ean,produto,valor FROM "+tabela
        self.cursor.execute(sql)
        contador = 0
        for linha in self.cursor.fetchall():
            contador += 1
            itens.append(linha)
        return itens

    def buscarProduto(self, termo):
        # print(termo)
        try:
            contador = 0
            sql = "SELECT * FROM produtos WHERE ean = ?"
            # termo = f'%{termo}%'
            self.cursor.execute(
                sql, (termo,))

            for linha in self.cursor.fetchall():
                contador += 1
                # print(linha)
                id = linha[0]
                ean = linha[1]
                produto = linha[2]
                quantidade = linha[3]
                valor_custo = linha[4]
                valor_venda = linha[5]
                ultima_compra = linha[6]

            return ean, produto, quantidade, valor_custo, valor_venda, ultima_compra, contador
        except Exception as e:
            return '', '', '', '', '', '', ''

    def cadastrarUsuario(self, ativo, login, senha, nome, cpf, funcao, botao):
        # print(botao)
        try:
            if ativo == 'SIM':
                ativo = '1'
            else:
                ativo = '2'

            ativo, login, senha, nome, cpf, funcao = str(ativo).upper(), login, senha, str(
                nome).upper(), str(cpf).upper(), str(funcao).upper()

            rativo, rlogin, rsenha, rnome, rcpf, rfuncao = self.buscarUsuario(
                termo=login)
            # print(botao)
            if str(botao) == 'Cadastrar':
                if rlogin == login or rcpf == cpf:
                    return False, False
                else:

                    sql = "INSERT INTO usuarios (ativo, login, senha, nome, cpf, funcao ) VALUES (?,?,?,?,?,?)"
                    self.cursor.execute(
                        sql, (ativo, login, senha, nome, cpf, funcao))
                    self.conn.commit()
                    return True, 'cadastrado'
            else:

                try:
                    sql = "UPDATE usuarios SET ativo=?, login=?, senha=?, nome=?, cpf=?,funcao=? WHERE login=?"
                    self.cursor.execute(
                        sql, (ativo, login, senha, nome, cpf, funcao, login))
                    self.conn.commit()
                    return True, 'atualizado'
                except Exception as e:
                    return False, False

        except Exception as e:
            # print(e)
            return False

    def editarUsuario(self, id, ativo, login, senha, nome, cpf, funcao):
        try:
            ativo, ean, produto, quantidade, valor_custo, valor_venda, ultima_compra = str(ativo).upper(), str(ean).upper(), str(
                produto).upper(), str(quantidade).upper(), str(valor_custo).upper().replace(',', '.'), str(valor_venda).upper().replace(',', '.'), str(ultima_compra).upper()
            sql = "UPDATE produtos SET ativo=?, ean=?, produto=?, quantidade=?, valor_custo=?, valor_venda=?, ultima_compra=? WHERE id=?"
            self.cursor.execute(sql, (ativo, ean, produto, quantidade,
                                valor_custo, valor_venda, ultima_compra, id))
            self.conn.commit()
            return f"O {produto=} {ean=} foi alterado com sucesso!"
        except Exception as e:
            return f"Ocorreu um erro ao alterar = {e}"

    def listar_tudoUsuario(self, tabela):
        itens = []
        sql = "SELECT * FROM "+tabela+" order by nome"
        self.cursor.execute(sql)
        contador = 0
        for linha in self.cursor.fetchall():
            contador += 1
            itens.append(linha)
        return itens

    def buscarUsuario(self, termo):
        # print(termo)
        try:
            contador = 0
            sql = "SELECT * FROM usuarios where login = ?"
            # termo = f'%{termo}%'
            self.cursor.execute(
                sql, (termo,))

            for linha in self.cursor.fetchall():

                contador += 1

                ativo = linha[0]
                login = linha[1]
                senha = linha[2]
                nome = linha[3]
                cpf = linha[4]
                funcao = linha[5]

            return ativo, login, senha, nome, cpf, funcao
        except Exception as e:
            return '', '', '', '', '', ''

    def geraGraficoMes(self, mes, ano, texto=False):
        try:
            test_date = datetime(ano, mes, 1)
            ultimo_dia_mes = calendar.monthrange(
                test_date.year, test_date.month)[1]

            # Executar a consulta SQL para obter as vendas diárias
            self.cursor.execute(
                "SELECT replace(data,'-','/'), total_compra FROM vendas where ativo = 'SIM'")
            dados_ativo = self.cursor.fetchall()

            self.cursor.execute(
                "SELECT replace(data,'-','/'), total_compra FROM vendas where ativo = 'CANCELADO'")
            dados_cancelado = self.cursor.fetchall()

            #################### ATIVO ########################################
            # cria um dataframe
            df = pd.DataFrame(dados_ativo, columns=['DATA', 'VENDA'])

            # corrige o valor
            df['VENDA'] = pd.to_numeric(df['VENDA'])

            # corrige a data
            df['DATA'] = df['DATA'].apply(lambda x: x[:10])
            df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
            # no need to wrap in Series
            s = pd.to_datetime(df['DATA'], unit='s')
            assert str(s.dtype) == 'datetime64[ns]'   # VERY IMPORTANT!!!!
            df.index = s
            df = df[f'{ano}-{mes}-01':f'{ano}-{mes}-{ultimo_dia_mes}']
            df.reset_index(drop=True, inplace=True)

            # agrupa as datas
            df = df.groupby('DATA').sum()

            # cria uma nova coluna
            df['DIA'] = df.index
            df['DIA'] = df['DIA'].astype(str).apply(lambda x: x[8:])
            resultado_ativo = df.groupby("DIA")['VENDA'].sum()

            #################### CANCELADO ########################################
            # cria um dataframe
            df = pd.DataFrame(dados_cancelado, columns=['DATA', 'CANCELADO'])

            # corrige o valor
            df['CANCELADO'] = pd.to_numeric(df['CANCELADO'])

            # corrige a data
            df['DATA'] = df['DATA'].apply(lambda x: x[:10])
            df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
            # no need to wrap in Series
            s = pd.to_datetime(df['DATA'], unit='s')
            assert str(s.dtype) == 'datetime64[ns]'   # VERY IMPORTANT!!!!
            df.index = s
            df = df[f'{ano}-{mes}-01':f'{ano}-{mes}-{ultimo_dia_mes}']
            df.reset_index(drop=True, inplace=True)

            # agrupa as datas
            df = df.groupby('DATA').sum()

            # cria uma nova coluna
            df['DIA'] = df.index
            df['DIA'] = df['DIA'].astype(str).apply(lambda x: x[8:])
            resultado_cancelado = df.groupby("DIA")['CANCELADO'].sum()

            df = pd.concat(
                [resultado_ativo, resultado_cancelado], axis=1)
            df['DIA'] = df.index

            ############# GRAFICO ##################################################

            fig, ax = plt.subplots()

            bar_width = 0.35
            x = np.arange(len(df['DIA']))
            bar1 = ax.bar(x - bar_width/2,
                          df['VENDA'], bar_width, label='VENDA')
            bar2 = ax.bar(x + bar_width/2,
                          df['CANCELADO'], bar_width, label='CANCELADO')

            # Adicionando labels com os valores a cada coluna
            for i, v in enumerate(df['VENDA']):
                ax.text(i - bar_width, v, f'R$ {v}', color='black',
                        va='bottom', fontweight='bold')
            for i, v in enumerate(df['CANCELADO']):
                ax.text(i, v, f'R$ {v}', color='black',
                        va='bottom')

            ax.set_xticks(x)
            ax.set_xticklabels(df['DIA'])
            ax.legend(loc='upper left')
            plt.title(f"Venda do mês {mes}/{ano}")
            plt.xlabel(f"Dia do mes {mes}")
            plt.ylabel("Valor Total do dia")
            # plt.savefig('grafico.jpg', dpi=200)
            if texto:
                plt.savefig('GraficoMes.jpg')
            else:
                plt.show()

            return df.loc[:, ['DIA', 'VENDA', 'CANCELADO']]

        except Exception as e:
            print(e)
            return False

    def geraGraficoAno(self, ano, texto=False):
        try:

            # Executar a consulta SQL para obter as vendas diárias
            self.cursor.execute(
                "SELECT replace(data,'-','/'), total_compra FROM vendas where ativo = 'SIM'")
            dados_ativo = self.cursor.fetchall()

            self.cursor.execute(
                "SELECT replace(data,'-','/'), total_compra FROM vendas where ativo = 'CANCELADO'")
            dados_cancelado = self.cursor.fetchall()

            #################### ATIVO ########################################
            # cria um dataframe
            df = pd.DataFrame(dados_ativo, columns=['DATA', 'VENDA'])
            # dfc = pd.DataFrame(dados_cancelado, columns=['DATA', 'CANCELADO'])

            # corrige o valor
            df['VENDA'] = pd.to_numeric(df['VENDA'])
            # dfc['VENDA'] = pd.to_numeric(dfc['CANCELADO'])

            # corrige a data
            df['DATA'] = df['DATA'].apply(lambda x: x[3:10])
            # dfc['DATA'] = dfc['CANCELADO'].apply(lambda x: x[3:10])

            df_final = pd.DataFrame()

            for i in range(1, 13):
                df_temp = df.loc[df['DATA'].isin([f'0{i}/{ano}'])]
                df_temp = df_temp.groupby('DATA').sum()
                if not df_temp.empty:
                    df_final = pd.concat([df_final, df_temp])

                # df_tempc = dfc.loc[dfc['DATA'].isin([f'0{i}/{ano}'])]
                # df_tempc = df_tempc.groupby('DATA').sum()
                # if not df_temp.empty:
                #     df_final = pd.concat([df_final, df_tempc])

            df_final['DATA'] = df_final.index

            # ############# GRAFICO ##################################################

            fig, ax = plt.subplots()

            bar_width = 0.35
            x = np.arange(len(df_final['DATA']))
            bar1 = ax.bar(x - bar_width/2,
                          df_final['VENDA'], bar_width, label='VENDA')
            # bar2 = ax.bar(x + bar_width/2,
            #   df['CANCELADO'], bar_width, label='CANCELADO')

            # Adicionando labels com os valores a cada coluna
            for i, v in enumerate(df_final['VENDA']):
                ax.text(i - bar_width, v, f'R$ {v}', color='black',
                        va='bottom', fontweight='bold')
            # for i, v in enumerate(df['CANCELADO']):
            #     ax.text(i, v, f'R$ {v}', color='red',
            #             va='bottom')

            ax.set_xticks(x)
            ax.set_xticklabels(df_final['DATA'])
            ax.legend('', frameon=False)
            plt.title(f"Venda de {ano}")
            plt.xlabel(f"Mes")
            plt.ylabel("Valor Total do mes")
            # plt.savefig('grafico.jpg', dpi=200)
            if texto:
                plt.savefig('GraficoAno.jpg')
            else:
                plt.show()

            return df_final
        except Exception as e:
            print(e)

    def geraGraficoTipo(self, mes, ano, texto=False):
        try:
            test_date = datetime(ano, mes, 1)
            ultimo_dia_mes = calendar.monthrange(
                test_date.year, test_date.month)[1]

            # Executar a consulta SQL para obter as vendas diárias
            self.cursor.execute(
                "SELECT replace(data,'-','/'), forma_pagamento FROM vendas where ativo = 'SIM'")
            dados_ativo = self.cursor.fetchall()

            #################### ATIVO ########################################
            # cria um dataframe
            df = pd.DataFrame(dados_ativo, columns=[
                              'DATA', 'FORMA_PAGAMENTO'])

            # corrige a data
            df['DATA'] = df['DATA'].apply(lambda x: x[:10])
            df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
            # no need to wrap in Series
            s = pd.to_datetime(df['DATA'], unit='s')
            assert str(s.dtype) == 'datetime64[ns]'   # VERY IMPORTANT!!!!
            df.index = s
            df = df[f'{ano}-{mes}-01':f'{ano}-{mes}-{ultimo_dia_mes}']
            df.reset_index(drop=True, inplace=True)

            # agrupa as datas
            df = df.groupby('FORMA_PAGAMENTO').count()

            # cria uma nova coluna
            df['DIA'] = df.index
            df['DIA'] = df['DIA'].astype(str).apply(lambda x: x[8:])

            ############# GRAFICO ##################################################

            # df.plot(y='DATA', kind='pie')
            df.plot.pie(y='DATA', figsize=(5, 5),
                        autopct='%1.1f%%', startangle=90)
            plt.legend('', frameon=False)
            plt.title(f"Venda por forma de pagamento no mês {mes}/{ano}")
            plt.xlabel(f"")
            plt.ylabel('')
            if texto:
                plt.savefig('GraficoTipo.jpg')
            else:
                plt.show()

            return df

        except Exception as e:
            print(e)
            return False

    def criaExcel(self, planilha):
        import openpyxl
        from datetime import date
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference, BarChart
        from openpyxl.styles import Font

        data_e_hora_atuais = datetime.now()
        data_atual = date.today()
        mes = data_e_hora_atuais.strftime('%m')
        ano = data_e_hora_atuais.strftime('%Y')

        # cria a planilha
        workbook = Workbook()
        workbook.save(planilha)

############################### resumo ############################################################################################
        aba = workbook.active
        aba.title = 'RESUMO'

        aba.column_dimensions['A'].width = 20
        aba.column_dimensions['B'].width = 20

        aba.insert_rows(1)
        aba.merge_cells('A1:B1')
        aba.cell(row=1, column=1).value = f'VENDA MES/MES DO ANO {ano}'
        aba[f"A2"] = 'MES REFERENCIA'
        aba[f"B2"] = 'VENDA R$'
        result = operacoes.geraGraficoAno(ano, texto=True)
        # gravar na planilha

############################### usuarios ############################################################################################
        aba = workbook.create_sheet(title="Usuarios")
        aba.insert_rows(1)

        aba[f"A1"] = 'ATIVO'
        aba[f"B1"] = 'LOGIN'
        aba[f"C1"] = 'NOME'
        aba[f"D1"] = 'CPF'
        aba[f"E1"] = 'FUNCAO'

        coluna = ['A', 'B', 'C', 'D', 'E']

        # negrito no titulo da coluna
        for i in range(len(coluna)):
            aba.cell(row=1, column=i+1).font = Font(bold=True)

        resultado = operacoes.listar_tudo(tabela='usuarios', tudo=True)

        # escreve os dados
        contador = 2
        for i in resultado:
            aba.insert_rows(contador)
            item = 0
            for x in coluna:
                if x == 'C':
                    item += 1
                    aba[f"{x}{contador}"] = i[item]
                else:
                    if x == 'A':
                        if i[item] == 1:
                            aba[f"{x}{contador}"] = 'ATIVO'
                        else:
                            aba[f"{x}{contador}"] = 'INATIVO'
                    else:
                        aba[f"{x}{contador}"] = i[item]
                item += 1
            contador += 1

        # dimensiona as colunas
        aba.auto_filter.ref = aba.dimensions
        aba.column_dimensions['B'].width = 20
        aba.column_dimensions['C'].width = 50
        aba.column_dimensions['D'].width = 15
        aba.column_dimensions['E'].width = 15
        aba.column_dimensions['F'].width = 15
        aba.column_dimensions['G'].width = 15


############################### produtos ############################################################################################
        aba = workbook.create_sheet(title="Estoque")
        aba.insert_rows(1)

        aba[f"A1"] = 'ATIVO'
        aba[f"B1"] = 'EAN'
        aba[f"C1"] = 'PRODUTO'
        aba[f"D1"] = 'ESTOQUE'
        aba[f"E1"] = 'PREÇO CUSTO'
        aba[f"F1"] = 'PREÇO VENDA'
        aba[f"G1"] = 'ULTIMA VENDA'
        coluna = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

        # negrito no titulo da coluna
        for i in range(len(coluna)):
            aba.cell(row=1, column=i+1).font = Font(bold=True)

        resultado = operacoes.listar_tudo(tabela='produtos', tudo=True)

        # escreve os dados
        contador = 2
        for i in resultado:
            aba.insert_rows(contador)
            item = 1
            for x in coluna:
                if x == 'A':
                    if i[item] == 1:
                        aba[f"{x}{contador}"] = 'ATIVO'
                    else:
                        aba[f"{x}{contador}"] = 'INATIVO'
                else:
                    aba[f"{x}{contador}"] = i[item]
                item += 1
            contador += 1

        # dimensiona as colunas
        aba.auto_filter.ref = aba.dimensions
        aba.column_dimensions['B'].width = 20
        aba.column_dimensions['C'].width = 50
        aba.column_dimensions['D'].width = 15
        aba.column_dimensions['E'].width = 15
        aba.column_dimensions['F'].width = 15
        aba.column_dimensions['G'].width = 15

 ############################### vendas ############################################################################################
        aba = workbook.create_sheet(title="Cupons")

        aba.insert_rows(1)

        aba[f"A1"] = 'COO'
        aba[f"B1"] = 'ATIVO'
        aba[f"C1"] = 'DATA'
        aba[f"D1"] = 'ITENS'
        aba[f"E1"] = 'TOTAL'
        aba[f"F1"] = 'FORMA PAGTO'
        aba[f"G1"] = 'VALOR PAGO'
        aba[f"H1"] = 'TROCO'
        aba[f"I1"] = 'OPERADOR'
        coluna = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

        # negrito no titulo da coluna
        for i in range(len(coluna)):
            aba.cell(row=1, column=i+1).font = Font(bold=True)

        resultado = operacoes.listar_tudo(tabela='vendas', tudo=True)

        # escreve os dados
        contador = 2
        for i in resultado:
            aba.insert_rows(contador)
            item = 0
            for x in coluna:
                aba[f"{x}{contador}"] = i[item]

                item += 1
            contador += 1

        # dimensiona as colunas
        aba.auto_filter.ref = aba.dimensions
        aba.column_dimensions['B'].width = 15
        aba.column_dimensions['C'].width = 20
        aba.column_dimensions['D'].width = 30
        aba.column_dimensions['E'].width = 15
        aba.column_dimensions['F'].width = 20
        aba.column_dimensions['G'].width = 20
        aba.column_dimensions['H'].width = 20
        aba.column_dimensions['I'].width = 20

        # grava a planilha
        workbook.save(filename=planilha)


if __name__ == "__main__":
    sistema = sys.platform
    if sistema == 'linux':
        operacoes = Operacoes('../DB/dbase.db')
        foto = 'img/tela.jpg'
    else:
        operacoes = Operacoes('..\\DB\\dbase.db')
        foto = 'img\\tela.jpg'

    print(operacoes.criaExcel('excel.xlsx'))
    # print(operacoes.geraGraficoAno(2023, texto=True))
    # print(operacoes.geraGraficoMes(2, 2023, texto=True))
    # print(operacoes.geraGraficoTipo(2, 2023, texto=True))
